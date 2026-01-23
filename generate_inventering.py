import os
import pandas as pd
import csv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Definiera filvägar med os.path.join
base_dir = os.path.dirname(os.path.abspath(__file__))
stock_file = os.path.join(base_dir, 'data', 'downloads', 'stock_report_2026-01-09.csv')
sales_file = os.path.join(base_dir, 'data', 'downloads', 'product_sales_items_2024-02-12_to_2026-01-09.csv')

# Läs in CSV-filerna med UTF-8 encoding
stock_df = pd.read_csv(stock_file, encoding='utf-8')
sales_df = pd.read_csv(sales_file, encoding='utf-8')

# Ta bort inledande punkt eller mellanslag i product_name
def clean_product_name(name):
    if isinstance(name, str) and (name.startswith('.') or name.startswith(' ')):
        return name[1:]
    return name

stock_df['product_name'] = stock_df['product_name'].apply(clean_product_name)
sales_df['product_name'] = sales_df['product_name'].apply(clean_product_name)

# Summera saldo per produkt över alla butiker (samma produkt kan finnas i flera butiker)
stock_clean = stock_df.groupby('product_name')['stock'].sum().reset_index()
stock_clean.columns = ['product_name', 'saldo']

# Beräkna genomsnittligt pris per produkt från sales (använd line_total)
sales_df['inköpspris'] = sales_df['line_total'] / 1.14 / 1.25

# Gruppera per produktnamn och beräkna genomsnittligt inköpspris samt hämta enhet
sales_prices = sales_df.groupby('product_name').agg({
    'inköpspris': 'mean',
    'unit': 'first'
}).reset_index()
sales_prices.columns = ['product_name', 'pris/enhet', 'enhet']

# Slå samman stock och sales data baserat på produktnamn
result_df = stock_clean.merge(sales_prices, on='product_name', how='left')

# Beräkna totalpris (saldo * pris/enhet)
result_df['totalpris'] = result_df['saldo'] * result_df['pris/enhet']

# Behåll produktnamn, saldo, enhet, inköpspris (pris/enhet) och totalpris
# Ordning: A=product_name, B=saldo, C=enhet, D=inköpspris, E=totalpris
final_df = result_df[['product_name', 'saldo', 'enhet', 'pris/enhet', 'totalpris']].copy()
final_df.columns = ['product_name', 'saldo', 'enhet', 'inköpspris', 'totalpris']

# Konvertera nummervärden till float och ersätt NaN-värden med 0.0
final_df['saldo'] = pd.to_numeric(final_df['saldo'], errors='coerce').fillna(0.0).astype(float)
final_df['inköpspris'] = pd.to_numeric(final_df['inköpspris'], errors='coerce').fillna(0.0).astype(float)
final_df['totalpris'] = pd.to_numeric(final_df['totalpris'], errors='coerce').fillna(0.0).astype(float)
final_df['enhet'] = final_df['enhet'].fillna('')

# Sortera så att de med inköpspris == 0 kommer först
final_df = final_df.sort_values(by='inköpspris', ascending=True).reset_index(drop=True)

# Spara resultatet med UTF-8-sig (med BOM) för bättre kompatibilitet med Excel
output_file_csv = os.path.join(base_dir, 'inventering_resultat.csv')
final_df.to_csv(output_file_csv, index=False, encoding='utf-8-sig', sep=';', quoting=csv.QUOTE_ALL)

# Skapa Excel-fil med formler i kolumn E (multiplicerar B och D)
output_file_xlsx = os.path.join(base_dir, 'inventering_resultat.xlsx')
wb = Workbook()
ws = wb.active

# Lägg till rubriker
# Ordning: A=product_name, B=saldo, C=enhet, D=inköpspris, E=totalpris
headers = ['product_name', 'saldo', 'enhet', 'inköpspris', 'totalpris']
ws.append(headers)

# Lägg till data (utan formler först)
# Alla nummervärden sparas som floats
for idx, row in final_df.iterrows():
    # Konvertera saldo och inköpspris till float (0.0 om tomt eller NaN)
    saldo_value = float(row['saldo']) if pd.notna(row['saldo']) else 0.0
    inköpspris_value = float(row['inköpspris']) if pd.notna(row['inköpspris']) else 0.0
    
    ws.append([
        row['product_name'],
        saldo_value,  # Float
        row['enhet'],
        inköpspris_value,  # Float
        None  # Kolumn E kommer att få en formel efter att tabellen skapats
    ])

# Konvertera till Excel Table FÖRST
# Detta gör att vi kan använda strukturreferenser i formeln
last_row = len(final_df) + 1
tab = Table(displayName="InventeringTable", ref=f"A1:E{last_row}")
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style
ws.add_table(tab)

# Lägg till formel med strukturreferenser för Excel Table
# Strukturreferenser säkerställer att formeln automatiskt kopieras när nya rader läggs till
# @[kolumnnamn] refererar till värdet i aktuell rad för den kolumnen
# Formeln hanterar floats - kontrollerar om värden är 0 eller tomma
if len(final_df) > 0:
    # Lägg till formel i första dataraden med strukturreferens
    # Excel kommer automatiskt kopiera denna formel när nya rader läggs till i tabellen
    # Formeln kontrollerar om saldo eller inköpspris är 0 eller tomt
    first_data_row = 2
    ws[f'E{first_data_row}'] = '=IF(OR([@saldo]=0,[@inköpspris]=0,[@saldo]="",[@inköpspris]=""),"",[@saldo]*[@inköpspris])'
    
    # Kopiera formeln till alla befintliga rader också
    for idx in range(1, len(final_df)):
        excel_row = idx + 2
        ws[f'E{excel_row}'] = '=IF(OR([@saldo]=0,[@inköpspris]=0,[@saldo]="",[@inköpspris]=""),"",[@saldo]*[@inköpspris])'

# Spara Excel-filen
wb.save(output_file_xlsx)

print(f"Inventering genererad! Resultat sparad till:")
print(f"  CSV: {output_file_csv}")
print(f"  Excel: {output_file_xlsx}")
print(f"Totalt antal produkter: {len(final_df)}")
print(f"\nFörsta 10 raderna:")
print(final_df.head(10))
