"""
Shared utilities for the Torp report generation pipeline.

This module centralises code that was previously duplicated across
2generate_picking_list.py, 3generate_order_list.py and
4generate_graphical_picking_list.py:

  - Path resolution (PROJECT_ROOT, parameter file lookup with _MEIPASS fallback)
  - Filename sanitisation (Windows-safe, handles all forbidden characters)
  - Write-permission helpers (clear read-only attribute, preflight write check,
    atomic temp+move with retry for fragile network shares)
  - Data loading helpers (find_latest_file, load_leveransfrekvens,
    load_and_prepare_sales_data)
  - Forecasting (predict_product_sales for total-period forecasts, plus the
    weekly-forecast helpers used by the graphical picking list)
"""

import os
import re
import sys
import stat
import glob
import time
import shutil
import subprocess
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

from sklearn.neighbors import KNeighborsRegressor

try:
    from sklearn.linear_model import LinearRegression, Ridge
    from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
    _ADVANCED_ML = True
except ImportError:
    _ADVANCED_ML = False

warnings.filterwarnings('ignore')


# ---------------------------------------------------------------------------
# Path resolution
# ---------------------------------------------------------------------------

def get_project_root():
    """
    Returns the project root directory.

    - When frozen (PyInstaller .exe): the directory that contains the .exe.
      This is the user-facing folder where data/, output/ and parameter files
      live next to the executable.
    - When running as a regular Python script: the parent of scripts/ (or
      the script's own directory if it isn't inside scripts/).
    """
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent.resolve()
    here = Path(__file__).parent.resolve()
    if here.name == 'scripts':
        return here.parent
    return here


def get_param_file_path(filename):
    """
    Locate a parameter file (e.g. Leveransfrekvens.csv).

    Search order:
      1. <project_root>/data/parametrar/<filename>  (user-editable, deployable)
      2. <_MEIPASS>/data/parametrar/<filename>     (bundled inside .exe, fallback)

    Returns the first existing path. If neither exists, returns option 1 so
    the caller can report a clear "file missing" error.
    """
    rel = Path('data') / 'parametrar' / filename
    primary = get_project_root() / rel
    if primary.exists():
        return primary
    if getattr(sys, 'frozen', False):
        meipass = getattr(sys, '_MEIPASS', None)
        if meipass:
            bundled = Path(meipass) / rel
            if bundled.exists():
                return bundled
    return primary


def get_format_file_path():
    """Path to data/parametrar/format.xlsx (user-editable product order/colors)."""
    return get_param_file_path('format.xlsx')


def get_format_write_path():
    """
    Writable format.xlsx next to the exe/script.

    Reads may fall back to the bundled copy inside the .exe (_MEIPASS), but
    updates must always land in the user's data/parametrar/ folder so Excel
    changes persist between runs.
    """
    return get_project_root() / 'data' / 'parametrar' / 'format.xlsx'


# ---------------------------------------------------------------------------
# Product format (format.xlsx) – order and text colour per product
# ---------------------------------------------------------------------------

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Border, Side
    from openpyxl.utils import get_column_letter as _oxl_col_letter
    _OPENPYXL_FORMAT = True
except ImportError:
    _OPENPYXL_FORMAT = False

_FORMAT_BORDER = None
_FORMAT_DEFAULT_COLOR = '000000'


def _format_border():
    global _FORMAT_BORDER
    if _FORMAT_BORDER is None and _OPENPYXL_FORMAT:
        thin = Side(style='thin')
        _FORMAT_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    return _FORMAT_BORDER


def _to_openpyxl_font_color(hex_rgb):
    """Convert 6-char RRGGBB to openpyxl aRGB ('FFAARRGGBB')."""
    h = str(hex_rgb or _FORMAT_DEFAULT_COLOR).replace('#', '').upper()
    if len(h) == 8:
        return h
    return 'FF' + h.zfill(6)[-6:]


def _cell_font_color_hex(cell):
    """Read font colour from an Excel cell as 6-char RRGGBB; default black."""
    if cell is None:
        return _FORMAT_DEFAULT_COLOR
    try:
        color = cell.font.color
    except AttributeError:
        return _FORMAT_DEFAULT_COLOR
    if color is None or getattr(color, 'type', None) in (None, 'default'):
        return _FORMAT_DEFAULT_COLOR
    if getattr(color, 'type', None) == 'rgb' and color.rgb:
        rgb = str(color.rgb).upper()
        if len(rgb) == 8:
            return rgb[2:]
        return rgb[-6:]
    return _FORMAT_DEFAULT_COLOR


class ProductFormatConfig:
    """Product order, name colours and qty-per-box from format.xlsx."""

    def __init__(self, entries):
        self.entries = entries
        self.order_index = {e['norm']: i for i, e in enumerate(entries)}
        self.color_by_norm = {e['norm']: e['color'] for e in entries}
        self.qty_per_box_by_norm = {
            e['norm']: e.get('qty_per_box') for e in entries
            if e.get('qty_per_box') not in (None, '')
        }


_FORMAT_HEADERS = ('Produktnamn', 'Produktkod', 'Mängd per låda')


def _format_header_row(name_cell):
    """True if the row looks like a column header rather than a product."""
    if name_cell is None or name_cell.value is None:
        return False
    label = str(name_cell.value).strip().lower()
    return label in ('produktnamn', 'product_name', 'namn', 'name')


def _parse_qty_per_box(value):
    """Parse mängd-per-låda from an Excel cell; return float/int or None."""
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        if float(value) == int(value):
            return int(value)
        return float(value)
    text = str(value).strip().replace(',', '.')
    if text in ('', 'nan', 'none'):
        return None
    try:
        num = float(text)
        if num == int(num):
            return int(num)
        return num
    except (ValueError, TypeError):
        return None


def _read_format_entries_from_workbook(path):
    """Load product rows from format.xlsx without modifying the file.

    Columns: A=product name (coloured), B=product code, C=mängd per låda
    (optional; user-maintained packing size).
    """
    if not _OPENPYXL_FORMAT:
        return []
    path = Path(path)
    if not path.exists():
        return []
    wb = load_workbook(str(path))
    ws = wb.active
    entries = []
    for row in ws.iter_rows(min_row=1, max_col=3):
        name_cell = row[0]
        code_cell = row[1] if len(row) > 1 else None
        qty_cell = row[2] if len(row) > 2 else None
        if name_cell.value is None or str(name_cell.value).strip() == '':
            continue
        if not entries and _format_header_row(name_cell):
            continue
        name = str(name_cell.value).strip()
        code_raw = code_cell.value if code_cell is not None else ''
        code = _normalise_product_code(code_raw) if code_raw not in (None, '') else ''
        entries.append({
            'name': name,
            'norm': normalize_name(name).lower(),
            'code': code,
            'color': _cell_font_color_hex(name_cell),
            'qty_per_box': _parse_qty_per_box(
                qty_cell.value if qty_cell is not None else None
            ),
        })
    wb.close()
    return entries


def _unique_products_from_stock(stock_df):
    """Unique (product_name, product_code) pairs from stock_report."""
    products = []
    seen = set()
    for _, row in stock_df.iterrows():
        name = row.get('product_name')
        if name is None or (isinstance(name, float) and pd.isna(name)):
            continue
        name = str(name).strip()
        if not name:
            continue
        norm = normalize_name(name).lower()
        if norm in seen:
            continue
        seen.add(norm)
        code = _normalise_product_code(row.get('product_code'))
        products.append((name, code))
    return products


def _format_workbook_needs_upgrade(path):
    """True if format.xlsx lacks header row and/or column C (mängd per låda)."""
    path = Path(path)
    if not path.exists() or not _OPENPYXL_FORMAT:
        return True
    wb = load_workbook(str(path))
    try:
        ws = wb.active
        if ws.max_column < 3:
            return True
        return not _format_header_row(ws.cell(1, 1))
    finally:
        wb.close()


def _write_format_workbook(path, entries, attempts=3, delay=2.0):
    """Write format.xlsx atomically (temp + replace) with retry on file locks."""
    if not _OPENPYXL_FORMAT:
        return
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.stem + '.tmp.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Format'
    border = _format_border()
    header_font = Font(bold=True, color=_to_openpyxl_font_color(_FORMAT_DEFAULT_COLOR))
    for col_idx, title in enumerate(_FORMAT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        if border is not None:
            cell.border = border
    for row_idx, entry in enumerate(entries, start=2):
        name_cell = ws.cell(row=row_idx, column=1, value=entry['name'])
        code_val = entry.get('code') or None
        code_cell = ws.cell(row=row_idx, column=2, value=code_val)
        qty_val = entry.get('qty_per_box')
        qty_cell = ws.cell(
            row=row_idx, column=3,
            value=qty_val if qty_val not in (None, '') else None,
        )
        color = entry.get('color', _FORMAT_DEFAULT_COLOR)
        name_cell.font = Font(color=_to_openpyxl_font_color(color))
        code_cell.font = Font(color=_to_openpyxl_font_color(_FORMAT_DEFAULT_COLOR))
        qty_cell.font = Font(color=_to_openpyxl_font_color(_FORMAT_DEFAULT_COLOR))
        if border is not None:
            name_cell.border = border
            code_cell.border = border
            qty_cell.border = border
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16

    last_err = None
    try:
        for i in range(attempts):
            try:
                if tmp.exists():
                    clear_readonly(tmp)
                wb.save(str(tmp))
                clear_readonly(tmp)
                if path.exists():
                    clear_readonly(path)
                os.replace(str(tmp), str(path))
                clear_readonly(path)
                return
            except OSError as e:
                last_err = e
                safe_unlink(tmp)
                if i < attempts - 1:
                    time.sleep(delay)
        if is_file_lock_error(last_err):
            raise PermissionError(format_write_lock_message(path, last_err)) from last_err
        raise last_err
    finally:
        wb.close()
        safe_unlink(tmp)


def load_and_sync_product_format(stock_df):
    """
    Read format.xlsx, append any stock_report products missing from the list
    (at the end, default black text), fill empty product codes from stock,
    save the workbook back, and return a ProductFormatConfig for sorting/colours.
    """
    read_path = get_format_file_path()
    write_path = get_format_write_path()
    print(f"\nLäser produktformat från: {read_path}")
    if read_path.resolve() != write_path.resolve():
        print(f"  Sparar uppdateringar till: {write_path}")

    entries = _read_format_entries_from_workbook(read_path)
    known_norms = {e['norm'] for e in entries}
    stock_products = _unique_products_from_stock(stock_df)
    stock_by_norm = {normalize_name(n).lower(): (n, c) for n, c in stock_products}

    if not entries and stock_products:
        for name, code in stock_products:
            norm = normalize_name(name).lower()
            entries.append({
                'name': name,
                'norm': norm,
                'code': code,
                'color': _FORMAT_DEFAULT_COLOR,
                'qty_per_box': None,
            })
        known_norms = {e['norm'] for e in entries}

    added = 0
    for norm, (name, code) in stock_by_norm.items():
        if norm not in known_norms:
            entries.append({
                'name': name,
                'norm': norm,
                'code': code,
                'color': _FORMAT_DEFAULT_COLOR,
                'qty_per_box': None,
            })
            known_norms.add(norm)
            added += 1

    codes_filled = 0
    for entry in entries:
        if 'qty_per_box' not in entry:
            entry['qty_per_box'] = None
        if not entry.get('code'):
            stock_match = stock_by_norm.get(entry['norm'])
            if stock_match and stock_match[1]:
                entry['code'] = stock_match[1]
                codes_filled += 1

    structure_upgrade = _format_workbook_needs_upgrade(write_path)
    if added > 0 or codes_filled > 0 or not write_path.exists() or structure_upgrade:
        try:
            _write_format_workbook(write_path, entries)
            if added > 0:
                print(
                    f"  [OK] Lade till {added} nya produkt(er) sist i format.xlsx "
                    f"(från stock_report)"
                )
            elif codes_filled > 0:
                print(
                    f"  [OK] Uppdaterade {codes_filled} produktkod(er) i format.xlsx"
                )
            elif structure_upgrade and write_path.exists():
                print(
                    f"  [OK] Uppdaterade format.xlsx med kolumnrubriker "
                    f"(Produktnamn / Produktkod / Mängd per låda)"
                )
            elif not write_path.exists():
                print(f"  [OK] Skapade format.xlsx med {len(entries)} produkter")
            else:
                print(f"  [OK] Uppdaterade format.xlsx ({len(entries)} produkter)")
        except OSError as e:
            hint = (
                format_write_lock_message(write_path, e)
                if is_file_lock_error(e) else str(e)
            )
            print(
                f"  [VARNING] Kunde inte uppdatera format.xlsx: {hint}\n"
                f"           Använder synkad konfiguration i minnet men filen "
                f"på disk är oförändrad."
            )
    else:
        print(f"  [OK] {len(entries)} produkter i format.xlsx (inga nya att lägga till)")

    return ProductFormatConfig(entries) if entries else None


def load_product_format():
    """Read-only load of format.xlsx (no sync). Returns None if missing/empty."""
    entries = _read_format_entries_from_workbook(get_format_file_path())
    return ProductFormatConfig(entries) if entries else None


def sort_dataframe_by_product_format(df, product_format, product_col='Produktnamn'):
    """Sort rows by format.xlsx order; unknown products go last (stable by name)."""
    if product_format is None or df is None or len(df) == 0:
        return df
    if product_col not in df.columns:
        return df
    fallback = len(product_format.entries) + 1000

    def _order_key(name):
        norm = normalize_name(name).lower()
        return product_format.order_index.get(norm, fallback)

    out = df.copy()
    out['_fmt_order'] = out[product_col].map(_order_key)
    out['_fmt_name'] = out[product_col].astype(str)
    out = out.sort_values(['_fmt_order', '_fmt_name'], kind='stable')
    return out.drop(columns=['_fmt_order', '_fmt_name'])


def sort_dataframe_by_store_and_format(df, product_format, store_col='store_name',
                                       product_col='Produktnamn'):
    """Sort each store block by format.xlsx product order."""
    if product_format is None or df is None or len(df) == 0:
        return df
    if store_col not in df.columns:
        return sort_dataframe_by_product_format(df, product_format, product_col)
    parts = []
    for _, group in df.groupby(store_col, sort=False):
        parts.append(sort_dataframe_by_product_format(group, product_format, product_col))
    return pd.concat(parts, ignore_index=True)


def write_formatted_excel(df, excel_path, sheet_name, product_format,
                          product_col='Produktnamn'):
    """
    Write DataFrame to Excel and apply product-name font colours from format.xlsx.
    """
    if not _OPENPYXL_FORMAT:
        return
    excel_path = Path(excel_path)
    if excel_path.exists():
        clear_readonly(excel_path)
    with pd.ExcelWriter(str(excel_path), engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        for idx, col in enumerate(df.columns, 1):
            max_length = max(df[col].astype(str).map(len).max(), len(str(col)))
            worksheet.column_dimensions[_oxl_col_letter(idx)].width = min(max_length + 2, 50)
        if product_format is not None and product_col in df.columns:
            col_idx = list(df.columns).index(product_col) + 1
            for row_offset, (_, row) in enumerate(df.iterrows(), start=2):
                norm = normalize_name(row[product_col]).lower()
                color = product_format.color_by_norm.get(norm, _FORMAT_DEFAULT_COLOR)
                cell = worksheet.cell(row=row_offset, column=col_idx)
                cell.font = Font(
                    color=_to_openpyxl_font_color(color),
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                )


# ---------------------------------------------------------------------------
# Filename safety
# ---------------------------------------------------------------------------

# Windows forbids these characters in filenames, plus control characters.
_FORBIDDEN_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def safe_filename(name):
    """
    Sanitise a string so it can be used as a Windows filename component.
    Replaces forbidden characters with underscore, normalises whitespace
    (NBSP, tabs, zero-width chars → regular space, collapsed) and strips
    trailing whitespace/dots, which Windows silently truncates.
    """
    if name is None:
        return '_'
    cleaned = _FORBIDDEN_FILENAME_CHARS.sub('_', str(name))
    cleaned = normalize_name(cleaned)
    cleaned = cleaned.rstrip('.')
    return cleaned or '_'


# ---------------------------------------------------------------------------
# Whitespace normalisation for names (store names, product names, etc.)
# ---------------------------------------------------------------------------
#
# A surprising number of bugs come from invisible whitespace differences in
# names that arrive via CSV exports from Excel, Word documents or copy-paste
# from emails. Examples we have to handle:
#
#   - Trailing/leading ASCII space:  "Sandsund kiosk " vs "Sandsund kiosk"
#   - Non-breaking space (U+00A0):   "Sandsund\u00a0kiosk" - looks identical
#     but != "Sandsund kiosk". Python's .strip() handles outer NBSP but
#     internal NBSP between words must be replaced.
#   - Multiple internal spaces:      "Sandsund  kiosk" vs "Sandsund kiosk"
#   - Tab characters:                "Sandsund\tkiosk"
#   - Zero-width characters:         U+200B (ZWSP), U+200C, U+200D, U+FEFF (BOM).
#     These are not in str.isspace() so .strip() does NOT remove them. They
#     appear invisibly when text is pasted from web pages or emails.
#
# normalize_name does all of this in one pass: remove zero-width characters,
# replace every Unicode whitespace run with a single ASCII space, and strip.

# Zero-width characters that survive .strip() and break exact matching.
_INVISIBLE_CHARS = re.compile(r'[\u200b-\u200d\ufeff]')
# Any run of whitespace (ASCII, NBSP, ideographic, etc.) collapses to one space.
_WHITESPACE_RUN = re.compile(r'\s+')


def normalize_name(value):
    """
    Normalise a single name/text value for reliable string matching.

    Returns the input unchanged for NaN/None. For other values: removes
    invisible characters, collapses any whitespace run to a single ASCII
    space, and strips outer whitespace.
    """
    if value is None:
        return value
    try:
        if pd.isna(value):
            return value
    except (TypeError, ValueError):
        pass
    s = str(value)
    s = _INVISIBLE_CHARS.sub('', s)
    s = _WHITESPACE_RUN.sub(' ', s).strip()
    return s


def normalize_name_series(series):
    """
    Vectorised normalize_name for a pandas Series. NaN is preserved.

    Uses regex .str operations which keep NaN automatically, much faster
    than .apply(normalize_name) on large frames.
    """
    s = series.astype('string')
    s = s.str.replace(_INVISIBLE_CHARS, '', regex=True)
    s = s.str.replace(_WHITESPACE_RUN, ' ', regex=True).str.strip()
    return s.astype(object)


def count_normalised_changes(original_series, normalised_series):
    """
    Count how many cells differ between the original and the normalised
    series. Used for diagnostic logging at data-ingestion points.
    """
    a = original_series.astype('string').fillna('')
    b = normalised_series.astype('string').fillna('')
    return int((a != b).sum())


# ---------------------------------------------------------------------------
# Write-permission helpers
# ---------------------------------------------------------------------------

def clear_readonly(path):
    """
    Ensure the file at `path` is writable by removing the read-only attribute.
    Silent no-op if the file does not exist.

    On Windows the FAT-style read-only attribute is the common reason a
    "Permission denied" error shows up when re-writing existing output files.
    NTFS ACLs are inherited from the parent directory and are not changed here.
    """
    p = Path(path)
    if not p.exists():
        return
    try:
        os.chmod(str(p), stat.S_IWRITE | stat.S_IREAD)
    except OSError:
        pass
    if os.name == 'nt':
        try:
            import subprocess
            flags = getattr(subprocess, 'CREATE_NO_WINDOW', 0)
            subprocess.run(
                ['attrib', '-R', str(p)],
                check=False,
                capture_output=True,
                creationflags=flags,
            )
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Inköpspris-logg
# ---------------------------------------------------------------------------
# Logg-format (JSON):
# {
#   "products": {
#     "<product_code>": {
#       "last_price": 12.45,
#       "last_seen": "2026-05-26",
#       "last_name": "Snellman Riktig Grillkorv 400 g",
#       "history": [
#         {"date": "2026-05-26", "price": 12.45, "name": "..."},
#         {"date": "2026-05-10", "price": 11.99, "name": "..."}
#       ]
#     }
#   }
# }
#
# Loggen är robust mot trasig/saknad fil - vid varje fel returneras ett
# tomt skelett så pipelinen kan fortsätta utan crash.

import json
from datetime import datetime as _datetime


def _empty_price_log():
    return {'products': {}}


def load_price_log(path):
    """
    Läs inköpsprisloggen från `path` (JSON). Vid alla typer av fel
    (filen saknas, JSON är trasig, fel format, fel encoding) returneras
    ett tomt skelett OCH en varning loggas - pipelinen ska aldrig krascha
    p.g.a. att loggen är skadad.
    """
    path = Path(path)
    if not path.exists():
        return _empty_price_log()
    try:
        with open(str(path), 'r', encoding='utf-8') as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError, UnicodeDecodeError) as e:
        print(
            f"  [VARNING] Kunde inte läsa inköpsprislogg {path}: {e}\n"
            f"           Skapar ny logg från noll - tidigare historik kan "
            f"vara förlorad men pipelinen kör vidare."
        )
        return _empty_price_log()
    if not isinstance(data, dict) or 'products' not in data \
            or not isinstance(data.get('products'), dict):
        print(
            f"  [VARNING] Inköpsprislogg {path} har oväntat format. "
            f"Skapar ny logg."
        )
        return _empty_price_log()
    return data


def save_price_log(log, path):
    """
    Skriv inköpsprisloggen atomiskt (skriv-till-temp + rename) så filen
    inte blir halvskriven om processen kraschar mitt under skrivning.
    Vid skrivfel skriver vi en varning men kraschar inte - loggen kommer
    bara att vara oförändrad nästa körning.
    """
    path = Path(path)
    assert_not_parameter_path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + '.tmp')
    try:
        if tmp.exists():
            clear_readonly(tmp)
        with open(str(tmp), 'w', encoding='utf-8') as f:
            json.dump(log, f, ensure_ascii=False, indent=2)
        if path.exists():
            clear_readonly(path)
        os.replace(str(tmp), str(path))
        clear_readonly(path)
    except OSError as e:
        print(
            f"  [VARNING] Kunde inte spara inköpsprislogg {path}: {e}\n"
            f"           Logikern fortsätter men historiken uppdateras inte."
        )
        try:
            tmp.unlink(missing_ok=True)
        except OSError:
            pass


def _normalise_product_code(code):
    """
    Standardisera product_code till en sträng vi kan använda som logg-nyckel.
    Tar bort .0 från floats (pandas-läst Excel-beteende) och whitespace.
    Returnerar None för NaN/tomt - sådana produkter loggas inte separat.
    """
    if code is None:
        return None
    try:
        if pd.isna(code):
            return None
    except (TypeError, ValueError):
        pass
    s = str(code).strip()
    if not s:
        return None
    try:
        # Numeriska koder som lagras som float ('6.409620e+12') → int-sträng.
        f = float(s)
        if not (f != f):  # not NaN
            i = int(f)
            if float(i) == f:
                return str(i)
    except (ValueError, TypeError):
        pass
    return s


def track_purchase_price(log, product_code, product_name, current_price,
                         today=None):
    """
    Uppdatera loggen för en produkt och returnera senaste kända pris att
    visa i orderlistan.

    - current_price är aktuellt pris från product_sales_items.unit_price
      (NaN/None om vi inte sett produkten i senaste items-fil).
    - Om current_price är giltig (>0 och inte NaN) och skiljer sig från
      senaste loggade pris → lägg till entry i history + uppdatera last_*.
    - Om current_price saknas/0 → fallback till loggens last_price.
    - Om varken aktuellt pris eller logg finns → returnera None.

    Muterar `log` in-place. Returnerar (returned_price, was_changed).
    """
    code_key = _normalise_product_code(product_code)
    if code_key is None:
        if current_price is not None:
            try:
                if pd.isna(current_price):
                    return None, False
                v = float(current_price)
                if v > 0:
                    return v, False
            except (TypeError, ValueError):
                pass
        return None, False

    entry = log.setdefault('products', {}).get(code_key)
    today_str = (today or _datetime.now()).strftime('%Y-%m-%d')

    current_valid = False
    current_value = None
    if current_price is not None:
        try:
            if not pd.isna(current_price):
                current_value = float(current_price)
                if current_value > 0:
                    current_valid = True
        except (TypeError, ValueError):
            pass

    if not current_valid:
        # Fallback till senaste loggade pris om vi inte ser produkten idag.
        if entry and entry.get('last_price') is not None:
            return float(entry['last_price']), False
        return None, False

    changed = False
    if entry is None:
        entry = {
            'last_price': current_value,
            'last_seen': today_str,
            'last_name': str(product_name) if product_name else None,
            'history': [{'date': today_str, 'price': current_value,
                        'name': str(product_name) if product_name else None}],
        }
        log['products'][code_key] = entry
        changed = True
    else:
        prev_price = entry.get('last_price')
        if prev_price is None or abs(float(prev_price) - current_value) > 1e-6:
            entry['history'] = (entry.get('history') or []) + [{
                'date': today_str,
                'price': current_value,
                'name': str(product_name) if product_name else None,
            }]
            entry['last_price'] = current_value
            entry['last_seen'] = today_str
            entry['last_name'] = str(product_name) if product_name else None
            changed = True
        else:
            # Pris oförändrat - uppdatera bara senast-sedd-datum.
            entry['last_seen'] = today_str

    return current_value, changed


def extract_latest_prices_from_items(items_df):
    """
    Bygg en mapping {product_code_normalised: latest_unit_price} från
    product_sales_items. Tar SENASTE created_at per product_code så
    vi alltid jämför med det aktuella priset, inte ett gammalt.
    Robust mot saknade kolumner och tom df.
    """
    if items_df is None or len(items_df) == 0:
        return {}
    required = {'product_code', 'unit_price'}
    if not required.issubset(set(items_df.columns)):
        return {}
    df = items_df.copy()
    if 'created_at' in df.columns:
        df['_sort_dt'] = pd.to_datetime(df['created_at'], errors='coerce')
        df = df.sort_values('_sort_dt', ascending=True)
    df = df.dropna(subset=['product_code', 'unit_price'])
    if len(df) == 0:
        return {}
    df['_key'] = df['product_code'].map(_normalise_product_code)
    df = df.dropna(subset=['_key'])
    # keep='last' efter sort_values('asc') = senaste created_at vinner
    df = df.drop_duplicates(subset=['_key'], keep='last')
    df['_price'] = pd.to_numeric(df['unit_price'], errors='coerce')
    df = df[df['_price'] > 0]
    return dict(zip(df['_key'], df['_price']))


PARAMETER_DIR_NAMES = {'parametrar', 'parameters'}


def _is_parameter_path(path):
    """
    Returnera True om `path` ligger inne i en `data/parametrar/`-katalog.
    Parameter-CSV:erna (Leveransfrekvens.csv, Beställningsfrekvens.csv) är
    konfigurationsfiler som BARA ska läsas - aldrig skrivas av scripten.
    """
    try:
        parts = Path(path).resolve().parts
    except (OSError, RuntimeError):
        parts = Path(path).parts
    parent_parts = parts[:-1]
    return any(p.lower() in PARAMETER_DIR_NAMES for p in parent_parts)


def assert_not_parameter_path(path):
    """
    Höj ett tydligt fel om något försöker skriva en parameterfil.
    Anropas av alla skrivhjälpare som första kontroll så att en framtida
    refactor inte råkar skriva över användarens Leveransfrekvens.csv eller
    Beställningsfrekvens.csv (vilket skulle radera deras konfiguration).
    """
    if _is_parameter_path(path):
        raise PermissionError(
            f"Förbjuden skrivning: {path} ligger under data/parametrar/. "
            f"Parameterfiler ska bara läsas av scripten, aldrig skrivas. "
            f"Om du verkligen vill ändra konfigurationen, redigera filen "
            f"manuellt i en editor."
        )


def safe_unlink(path):
    """
    Delete a file if it exists. Returns True if a file was actually removed,
    False if the path didn't exist or wasn't a regular file.

    Never raises FileNotFoundError - "already gone" is considered success
    (that's the desired end state anyway). Other OSErrors (e.g. permission
    denied because someone has the file open) are caught and logged as a
    warning, returning False, so the calling pipeline keeps running.
    """
    p = Path(path)
    if not p.exists():
        # Already gone - that's success in the "be content" sense.
        return False
    try:
        # missing_ok=True so a race condition between the exists() check
        # above and the unlink() call below still doesn't crash.
        p.unlink(missing_ok=True)
        return True
    except FileNotFoundError:
        return False
    except OSError as e:
        # File is open in another process, ACL denies delete, etc.
        # Don't crash - the caller almost always just wants the file gone
        # and can move on if we can't manage it.
        print(f"  [WARNING] Kunde inte radera {p}: {e}")
        return False


def filter_garbage_product_names(stock_df, min_length=2):
    """
    Ta bort produkter med uppenbart skräpiga namn ur stock_report:
    - tomma strängar / endast whitespace
    - namn kortare än `min_length` tecken efter strip()
      (Användarens leverantörssystem har t.ex. en produkt som heter bara 'R'
      som inte ska följa med ner i plocklista/orderlista.)

    Robust mot saknade kolumner och NaN. Loggar vad som filtrerats bort.
    """
    if stock_df is None or len(stock_df) == 0 or 'product_name' not in stock_df.columns:
        return stock_df
    names = stock_df['product_name'].astype('string').fillna('').str.strip()
    too_short = names.str.len() < min_length
    if int(too_short.sum()) == 0:
        return stock_df
    dropped = stock_df.loc[too_short, ['store_name', 'product_name']]
    print(f"  Filtrerade bort {int(too_short.sum())} produkter med ogiltigt/för kort namn:")
    for _, row in dropped.drop_duplicates().head(10).iterrows():
        print(f"    - {row['store_name']!r}: {row['product_name']!r}")
    if len(dropped.drop_duplicates()) > 10:
        print(f"    ... och {len(dropped.drop_duplicates()) - 10} till")
    return stock_df.loc[~too_short].copy()


def deduplicate_stock(stock_df):
    """
    Stock-rapporten kan ha flera rader för samma (butik, produktnamn) när
    butiken har olika varianter i sitt POS (t.ex. 'Kållby Mugg' med tre
    olika product_id, eller barcode-NaN vs barcode-med-värde). Det gör att
    plocklistan kan visa samma produkt två gånger.

    Här slår vi ihop sådana rader per (butik, normaliserat-produktnamn):
    - `stock` summeras (om samma produkt ligger i tre varianter, är totala
      lagret summan av dem - det är vad användaren ser i butiken).
    - `stock_warning_limit` tas som SUMMA också, eftersom varje variant har
      sin egen varningsgräns och totalbehovet är summan.
    - Övriga kolumner: behåller den första (godtyckligt men deterministiskt).

    Loggar hur många rader som slogs samman.
    """
    if stock_df is None or len(stock_df) == 0:
        return stock_df
    if 'store_name' not in stock_df.columns or 'product_name' not in stock_df.columns:
        return stock_df

    key_cols = ['store_name', 'product_name']
    before = len(stock_df)
    dup_mask = stock_df.duplicated(subset=key_cols, keep=False)
    n_dup_groups = int(stock_df.loc[dup_mask, key_cols].drop_duplicates().shape[0])
    if not dup_mask.any():
        return stock_df

    numeric_cols = [c for c in ('stock', 'stock_warning_limit')
                    if c in stock_df.columns]
    other_cols = [c for c in stock_df.columns
                  if c not in key_cols + numeric_cols]

    agg = {c: 'sum' for c in numeric_cols}
    for c in other_cols:
        agg[c] = 'first'

    merged = (
        stock_df
        .groupby(key_cols, as_index=False, sort=False, dropna=False)
        .agg(agg)
    )
    merged = merged[stock_df.columns.tolist()]
    rows_collapsed = before - len(merged)
    print(
        f"  Slog ihop {rows_collapsed} duplicerade stock-rader till "
        f"{n_dup_groups} unika (butik, produkt)-kombinationer "
        f"(summerade stock + warning_limit)"
    )
    return merged


def filter_active_products(stock_df):
    """
    Behåll endast produkter där `product_status == 1` (aktiva) i stock_report,
    så att inaktiva produkter försvinner ur hela pipelinen.

    Robust mot olika varianter av stock_report-schema:
    - Saknas kolumnen helt (gammal datafil)? → returnera df oförändrad +
      skriv en varning. Vi vill inte krascha bara för att källschemat
      ändras igen i framtiden.
    - Finns kolumnen två gånger (`product_status` + `product_status.1`,
      vilket pandas auto-döper duplikatkolumner till)? → behåll den första,
      droppa duplikaten. Den nuvarande CSV-export från source-systemet har
      av någon anledning båda, men de innehåller identisk data.
    - Värdet är inte exakt 1 (t.ex. NaN, sträng, annat nummer)? → tolkas
      som "inte aktiv" och filtreras bort. Säkrare att felaktigt droppa
      än att felaktigt ta med en inaktiv produkt.

    Antas köras EFTER att df läst in CSV:n och INNAN build_stock_index,
    så att hela övriga pipelinen (som redan filtrerar mot stock_df) får
    den nya begränsningen gratis.
    """
    if stock_df is None or len(stock_df) == 0:
        return stock_df

    if 'product_status' not in stock_df.columns:
        print(
            "  [VARNING] Ingen 'product_status'-kolumn i stock_report - "
            "kan inte filtrera bort inaktiva produkter. Kör vidare som "
            "om alla produkter vore aktiva."
        )
        return stock_df

    duplicates = [c for c in stock_df.columns if c.startswith('product_status.')]
    if duplicates:
        stock_df = stock_df.drop(columns=duplicates)

    status = pd.to_numeric(stock_df['product_status'], errors='coerce')
    active_mask = status == 1
    inactive_count = int((~active_mask).sum())
    if inactive_count > 0:
        print(
            f"  Filtrerade bort {inactive_count} inaktiva produkter "
            f"(product_status != 1) ur stock_report"
        )
    return stock_df[active_mask].copy()


def preflight_writable(directory):
    """
    Verify that `directory` exists and we can write into it. Raises
    PermissionError with a clear message if not. Used to fail fast at the
    start of a step instead of blowing up after minutes of work.
    """
    d = Path(directory)
    d.mkdir(parents=True, exist_ok=True)
    probe = d / f'.write_test_{os.getpid()}_{int(time.time() * 1000)}'
    try:
        probe.write_text('ok', encoding='utf-8')
    except Exception as e:
        raise PermissionError(
            f"Output-katalogen är inte skrivbar: {d}\n"
            f"Ursprungligt fel: {e}\n"
            f"Kontrollera NTFS-rättigheter och att ingen annan process håller filer öppna."
        ) from e
    safe_unlink(probe)


def atomic_copy_with_retry(src, dst, attempts=3, delay=2.0):
    r"""
    Move `src` to `dst`, retrying on transient PermissionError.

    Intended for moving a file produced locally to a network share
    (\\server\share\...). SMB shares, antivirus scanning and PDF readers
    sometimes hold a brief lock on the destination; retrying smooths over
    those transients without losing the produced file.

    Side effects:
      - Creates dst.parent if it doesn't exist.
      - Clears the read-only attribute on dst (before and after copy).
      - Removes src on success.
    """
    src = Path(src)
    dst = Path(dst)
    dst.parent.mkdir(parents=True, exist_ok=True)

    last_err = None
    for i in range(attempts):
        try:
            if dst.exists():
                clear_readonly(dst)
            shutil.copy2(str(src), str(dst))
            safe_unlink(src)
            clear_readonly(dst)
            return
        except OSError as e:
            # PermissionError ärver från OSError - täcker både låsta filer
            # och transienta SMB/NFS-fel.
            last_err = e
            if i < attempts - 1:
                time.sleep(delay)

    if is_file_lock_error(last_err):
        raise PermissionError(format_write_lock_message(dst, last_err)) from last_err
    raise last_err


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------

def find_latest_file(pattern, directory, min_size_bytes=0):
    """
    Find the most recent file in `directory` matching `pattern` (e.g.
    'product_sales_*.csv'), using the latest YYYY-MM-DD found in the
    filename as the sort key. Raises FileNotFoundError if nothing matches.

    Smart special-case: `product_sales_*.csv` deliberately excludes
    `product_sales_items_*.csv` files. The two file *types* share the
    prefix `product_sales_` but represent very different data (daily
    snapshot vs full item history), and a naive glob would let the
    items file win the "latest" race because its name contains a more
    recent date in the to-DATE part.

    `min_size_bytes` lets callers skip files that are too small to be
    real data (e.g. header-only CSV exports of ~200 bytes). This protects
    against the SFTP export occasionally producing an empty cumulative
    items file for the current day; we then fall back to the previous
    day's file instead of silently loading "0 rows" downstream.
    """
    directory = Path(directory)
    if not directory.exists():
        raise FileNotFoundError(f"Katalogen finns inte: {directory}")

    files = glob.glob(str(directory / pattern))
    if pattern == 'product_sales_*.csv':
        files = [f for f in files
                 if not os.path.basename(f).startswith('product_sales_items_')]
    if not files:
        existing = list(directory.glob('*.csv'))
        if existing:
            print(
                f"  Hittade {len(existing)} CSV-filer i {directory}, "
                f"men inga matchar mönstret {pattern}"
            )
            print(f"  Exempel filer: {[f.name for f in existing[:5]]}")
        else:
            print(f"  Katalogen {directory} är tom eller innehåller inga CSV-filer")
        raise FileNotFoundError(f"Inga filer matchar mönstret {pattern} i {directory}")

    if min_size_bytes > 0:
        too_small = []
        kept = []
        for f in files:
            try:
                size = os.path.getsize(f)
            except OSError:
                size = 0
            if size < min_size_bytes:
                too_small.append((os.path.basename(f), size))
            else:
                kept.append(f)
        if too_small:
            print(
                f"  Hoppar över {len(too_small)} fil(er) under {min_size_bytes} "
                f"byte (sannolikt tomma/header-bara exporter):"
            )
            for name, size in too_small[:5]:
                print(f"    - {name} ({size} byte)")
            if len(too_small) > 5:
                print(f"    ... och {len(too_small) - 5} till")
        if not kept:
            raise FileNotFoundError(
                f"Alla filer som matchar {pattern} i {directory} är mindre än "
                f"{min_size_bytes} byte - inga giltiga datafiler att läsa."
            )
        files = kept

    file_dates = []
    for file in files:
        all_dates = re.findall(r'(\d{4}-\d{2}-\d{2})', os.path.basename(file))
        if all_dates:
            file_dates.append((max(all_dates), file))

    if not file_dates:
        raise ValueError(f"Kunde inte hitta datum i filnamn för {pattern}")

    file_dates.sort(key=lambda x: x[0], reverse=True)
    latest_file = file_dates[0][1]
    print(
        f"  Hittade senaste fil: {os.path.basename(latest_file)} "
        f"(datum: {file_dates[0][0]})"
    )
    return latest_file


def find_item_metadata_file(directory, min_size_bytes=1024):
    """
    Return the best file to read item-level metadata (unit, supplier,
    product code, prices) from.

    Prefers the cumulative `product_sales_items_*.csv` export, but the SFTP
    server intermittently produces a header-only (~200 byte) items file for
    the current period. In that case we DON'T abort the whole run: the daily
    `product_sales_*.csv` snapshots carry the exact same item-level schema
    (`order_status`, `store_name`, `product_name`, `unit`, `supplier_code`,
    `supplier_name`, `product_code`, `unit_price`, ...), so we fall back to
    the latest daily snapshot. This mirrors the fallback that
    `load_sales_history` already does for the forecasting history, so the
    mapping/price lookups stay consistent with it instead of crashing.

    Raises FileNotFoundError only if there is neither a usable items file
    nor any daily snapshot to fall back to.
    """
    directory = Path(directory)
    try:
        return find_latest_file(
            'product_sales_items_*.csv', directory, min_size_bytes=min_size_bytes
        )
    except FileNotFoundError as items_err:
        print(
            f"  [INFO] Ingen användbar items-fil "
            f"(product_sales_items_*.csv >= {min_size_bytes} byte): {items_err}"
        )
        print(
            "  Faller tillbaka på senaste dagliga product_sales_*.csv "
            "(samma item-schema) för enhets-/leverantörs-/prisinformation."
        )
        try:
            return find_latest_file('product_sales_*.csv', directory)
        except FileNotFoundError as daily_err:
            raise FileNotFoundError(
                "Hittade varken en användbar product_sales_items_*.csv "
                f"(>= {min_size_bytes} byte) eller någon daglig "
                f"product_sales_*.csv i {directory}. "
                f"Items-fel: {items_err} | Daglig-fel: {daily_err}"
            ) from daily_err


# ---------------------------------------------------------------------------
# Parameter CSV reading (encoding fallback for Excel-saved files on Windows)
# ---------------------------------------------------------------------------

_PARAM_CSV_ENCODINGS = ('utf-8-sig', 'cp1252')


def read_text_with_encoding_fallback(file_path, encodings=_PARAM_CSV_ENCODINGS):
    """
    Read a text file, trying UTF-8 (with BOM) first then Windows-1252.

    Excel on Swedish Windows often saves CSV as cp1252; our pipeline prefers
    UTF-8. Logs an info line when the fallback encoding is used.
    """
    file_path = Path(file_path)
    last_err = None
    for i, enc in enumerate(encodings):
        try:
            text = file_path.read_text(encoding=enc)
            if i > 0:
                print(
                    f"  Info: Läste {file_path.name} med {enc}-kodning "
                    f"(filen är inte UTF-8; spara som CSV UTF-8 i Excel för bästa resultat)."
                )
            return text
        except UnicodeDecodeError as e:
            last_err = e
    raise last_err


def read_csv_with_encoding_fallback(file_path, **read_csv_kwargs):
    """pd.read_csv wrapper with utf-8-sig → cp1252 fallback for parameter files."""
    file_path = Path(file_path)
    last_err = None
    for i, enc in enumerate(_PARAM_CSV_ENCODINGS):
        try:
            df = pd.read_csv(str(file_path), encoding=enc, **read_csv_kwargs)
            if i > 0:
                print(
                    f"  Info: Läste {file_path.name} med {enc}-kodning "
                    f"(filen är inte UTF-8; spara som CSV UTF-8 i Excel för bästa resultat)."
                )
            return df
        except UnicodeDecodeError as e:
            last_err = e
    raise last_err


def is_file_lock_error(err):
    """True when an OSError likely means the destination file is open/locked."""
    return isinstance(err, PermissionError) or (
        isinstance(err, OSError) and getattr(err, 'errno', None) == 13
    )


def format_write_lock_message(path, err):
    """
    Human-readable hint when writing to output fails because the file is locked.
    """
    path = Path(path)
    ext = path.suffix.lower()
    if ext in ('.xlsx', '.xls'):
        app = 'Excel'
    elif ext == '.pdf':
        app = 'en PDF-läsare'
    else:
        app = 'ett annat program'
    return (
        f"Kunde inte skriva till {path} – filen är troligen öppen i {app}. "
        f"Stäng filen och kör om (eller undvik att ha output-mappen öppen "
        f"under körning). Ursprungligt fel: {err}"
    )


def reraise_if_write_locked(path, err):
    """Re-raise with a clearer message when a write failed due to a file lock."""
    if is_file_lock_error(err):
        raise PermissionError(format_write_lock_message(path, err)) from err
    raise err


def preflight_file_not_locked(path):
    """
    Raise PermissionError if `path` exists but appears locked by another process
    (typically Excel holding an .xlsx open on Windows).
    """
    path = Path(path)
    if not path.exists():
        return
    clear_readonly(path)
    try:
        with open(path, 'a'):
            pass
    except OSError as e:
        reraise_if_write_locked(path, e)


def preflight_torp_excel_files(project_root=None):
    """
    Fail fast when Torp Excel files the pipeline may overwrite are locked.

    Checks format.xlsx and any existing .xlsx under output/plocklistor,
    output/orderlistor and output/orderlistor_leverantor. Raises
    PermissionError with a list of locked paths so the user can close Excel
    and re-run instead of failing mid-pipeline.
    """
    root = Path(project_root) if project_root is not None else get_project_root()
    candidates = [root / 'data' / 'parametrar' / 'format.xlsx']
    for rel in (
        'output/plocklistor',
        'output/orderlistor',
        'output/orderlistor_leverantor',
    ):
        folder = root / rel
        if folder.is_dir():
            candidates.extend(sorted(folder.glob('*.xlsx')))

    seen = set()
    locked = []
    for path in candidates:
        resolved = Path(path).resolve()
        if resolved in seen or not resolved.exists():
            continue
        seen.add(resolved)
        try:
            preflight_file_not_locked(resolved)
        except PermissionError:
            locked.append(resolved)

    if not locked:
        return

    lines = [
        'AVBRYTER: Följande Excel-filer är låsta (stäng dem i Excel och kör om):',
    ]
    lines.extend(f'  - {p}' for p in locked)
    raise PermissionError('\n'.join(lines))


def close_all_excel_applications():
    """
    Stäng alla Excel-processer på Windows så att låsta .xlsx-filer kan
    skrivas om. Osparade ändringar i öppna arbetsböcker går förlorade.

    Ordning:
      1. Försök stänga snyggt via COM (win32com) om Excel svarar.
      2. Force-kill EXCEL.EXE via taskkill som fallback.
    """
    if sys.platform != 'win32':
        print("  [INFO] Hoppar över Excel-stängning (inte Windows).")
        return

    closed_via_com = False
    try:
        import win32com.client  # type: ignore

        try:
            excel = win32com.client.GetActiveObject('Excel.Application')
        except Exception:
            excel = None
        if excel is not None:
            try:
                excel.DisplayAlerts = False
                # Workbooks-collection shrinks as we close; iterate a snapshot.
                count = int(excel.Workbooks.Count)
                for i in range(count, 0, -1):
                    try:
                        excel.Workbooks(i).Close(SaveChanges=False)
                    except Exception:
                        pass
                excel.Quit()
                closed_via_com = True
                print(f"  [OK] Stängde Excel via COM ({count} arbetsbok(er)).")
            except Exception as e:
                print(f"  [VARNING] COM-stängning av Excel misslyckades: {e}")
    except ImportError:
        pass
    except Exception as e:
        print(f"  [VARNING] Kunde inte anropa Excel via COM: {e}")

    try:
        result = subprocess.run(
            ['taskkill', '/F', '/IM', 'EXCEL.EXE'],
            capture_output=True,
            text=True,
            check=False,
        )
        # 0 = killed something, 128 = process not found (already closed).
        if result.returncode == 0:
            print("  [OK] Force-stängde kvarvarande EXCEL.EXE-process(er).")
        elif result.returncode == 128:
            if not closed_via_com:
                print("  [OK] Ingen Excel-process körde.")
        else:
            stderr = (result.stderr or '').strip()
            if stderr:
                print(f"  [VARNING] taskkill Excel: {stderr}")
    except OSError as e:
        print(f"  [VARNING] Kunde inte köra taskkill för Excel: {e}")

    # Ge Windows en kort stund att släppa filhandtag.
    time.sleep(1.0)


# ---------------------------------------------------------------------------
# Leveransfrekvens (delivery frequency per store)
# ---------------------------------------------------------------------------

def load_leveransfrekvens(file_path):
    """
    Read Leveransfrekvens.csv → dict {store_name: delivery_frequency_days}.
    Returns an empty dict and logs a warning if the file is missing or
    malformed; callers fall back to a default frequency in that case.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        print(f"Varning: Leveransfrekvens-fil saknas: {file_path}")
        print("  Använder standardvärde: 7 dagar för alla butiker")
        return {}

    try:
        df = read_csv_with_encoding_fallback(file_path, sep=';')
        print(f"  Hittade kolumner: {list(df.columns)}")

        store_col = None
        freq_col = None
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if 'store' in col_lower or 'butik' in col_lower:
                store_col = col
            if 'leveransfrekvens' in col_lower or 'frekvens' in col_lower:
                freq_col = col

        if store_col is None or freq_col is None:
            if len(df.columns) >= 2:
                store_col, freq_col = df.columns[0], df.columns[1]
                print(
                    f"  Varning: Kunde inte hitta kolumnnamn, "
                    f"använder: '{store_col}' och '{freq_col}'"
                )
            else:
                print("  Fel: CSV-filen har för få kolumner. Förväntar minst 2 kolumner.")
                return {}

        parametrar = {}
        for _, row in df.iterrows():
            butik_raw = row[store_col]
            if not pd.notna(butik_raw):
                continue
            # normalize_name (not just .strip()) so a stray NBSP or double
            # space in the CSV resolves to the same key as the sales data.
            butik = normalize_name(butik_raw)
            if not butik:
                continue
            try:
                freq_raw = row[freq_col]
                if pd.notna(freq_raw):
                    freq_str = str(freq_raw).strip()
                    freq_clean = ''.join(c for c in freq_str if c.isdigit() or c == '.')
                    if freq_clean:
                        parametrar[butik] = int(float(freq_clean))
            except (ValueError, TypeError):
                print(f"  Varning: Kunde inte tolka frekvens för '{butik}': {freq_raw}")

        print(f"Laddade leveransfrekvenser för {len(parametrar)} butiker")
        if parametrar:
            print(f"  Exempel: {list(parametrar.items())[:3]}")
        return parametrar
    except Exception as e:
        print(f"Varning: Kunde inte ladda leveransfrekvenser: {e}")
        print("  Använder standardvärde: 7 dagar för alla butiker")
        return {}


# ---------------------------------------------------------------------------
# Sales data loading
# ---------------------------------------------------------------------------

def _prepare_sales_df(df, keep_unit=False):
    """
    Normalise a raw sales DataFrame (read from product_sales_*.csv or
    product_sales_items_*.csv) into the pipeline's canonical schema:
    date, store, name, quantity (+ unit if keep_unit=True).

    Extracted from load_and_prepare_sales_data so the same standardisation
    can be reused when we concatenate several files in
    load_sales_history (which dedupes raw rows by order_item_id before
    handing the merged DataFrame back).
    """
    if 'period' in df.columns:
        df = df.rename(columns={
            'period': 'date',
            'store_name': 'store',
            'product_name': 'name',
            'total_quantity_sold': 'quantity',
            'total_sales': 'line_price',
        })
        df['date'] = pd.to_datetime(df['date']).dt.date
        if keep_unit and 'unit' not in df.columns:
            df['unit'] = 'st'
    elif 'created_at' in df.columns:
        df = df.rename(columns={
            'created_at': 'updated',
            'store_name': 'store',
            'product_name': 'name',
            'line_total': 'line_price',
        })
        df['updated'] = pd.to_datetime(df['updated'])
        df['date'] = df['updated'].dt.date
        if keep_unit and 'unit' not in df.columns:
            df['unit'] = 'st'
    else:
        date_cols = [c for c in df.columns
                     if 'date' in c.lower() or 'created' in c.lower() or 'period' in c.lower()]
        if date_cols:
            df['date'] = pd.to_datetime(df[date_cols[0]]).dt.date
        else:
            raise ValueError(
                f"Kunde inte hitta datumkolumn i filen. "
                f"Tillgängliga kolumner: {list(df.columns)}"
            )
        if 'store_name' in df.columns:
            df['store'] = df['store_name']
        if 'product_name' in df.columns:
            df['name'] = df['product_name']
        if 'quantity' not in df.columns:
            qty_cols = [c for c in df.columns if 'quantity' in c.lower()]
            if qty_cols:
                df['quantity'] = df[qty_cols[0]]
            else:
                raise ValueError(
                    f"Kunde inte hitta quantity-kolumn i filen. "
                    f"Tillgängliga kolumner: {list(df.columns)}"
                )

    if keep_unit and 'unit' not in df.columns:
        df['unit'] = 'st'

    required = ['date', 'store', 'name', 'quantity']
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"Saknade kolumner: {missing}. Tillgängliga kolumner: {list(df.columns)}"
        )

    if 'order_status' in df.columns:
        before = len(df)
        df = df[df['order_status'] == 'complete'].copy()
        if before != len(df):
            print(
                f"  Filtrerade bort {before - len(df)} rader med "
                f"order_status != 'complete'"
            )

    if 'quantity' in df.columns:
        nan_count = df['quantity'].isna().sum()
        if nan_count:
            print(f"  Fyller {nan_count} NaN-värden i quantity med 0")
            df['quantity'] = df['quantity'].fillna(0)

    if keep_unit and 'unit' in df.columns:
        nan_count = df['unit'].isna().sum()
        if nan_count:
            df['unit'] = df['unit'].fillna('st')

    before = len(df)
    df = df.dropna(subset=['name', 'store'])
    if before != len(df):
        print(f"  Removed {before - len(df)} rows with missing product name or store")

    # Normalise whitespace in name/store columns so that small differences
    # in CSV exports (trailing space, NBSP, double space, BOM) don't cause
    # "Sandsund kiosk" and "Sandsund kiosk " to be treated as separate stores.
    raw_store = df['store'].astype('string')
    raw_name = df['name'].astype('string')
    df['store'] = normalize_name_series(df['store'])
    df['name'] = normalize_name_series(df['name'])

    store_changed = count_normalised_changes(raw_store, df['store'])
    name_changed = count_normalised_changes(raw_name, df['name'])
    if store_changed or name_changed:
        print(
            f"  Normaliserade whitespace i butiks-/produktnamn: "
            f"{store_changed} butiksnamn och {name_changed} produktnamn justerade"
        )

    if keep_unit and 'unit' in df.columns:
        df['unit'] = normalize_name_series(df['unit'])

    print(f"  Laddat {len(df)} rader")
    print(f"  Butiker: {df['store'].nunique()}")
    print(f"  Produkter: {df['name'].nunique()}")
    print(f"  Datumintervall: {df['date'].min()} till {df['date'].max()}")

    return df


def load_and_prepare_sales_data(file_path, keep_unit=False):
    """
    Read and normalise product sales data from product_sales_*.csv or the
    transactional product_sales_items_*.csv layout. Returns a DataFrame with
    columns: date, store, name, quantity (+ unit if keep_unit=True).
    """
    file_path = Path(file_path)
    print(f"Läser försäljningsdata från {file_path}...")
    df = pd.read_csv(str(file_path))
    return _prepare_sales_df(df, keep_unit=keep_unit)


def load_sales_history(directory, keep_unit=False, min_items_size_bytes=1024):
    """
    Build the most complete sales-history DataFrame available by combining:

      1. The latest non-empty `product_sales_items_*.csv` (cumulative file
         going back to 2024-02-12; this is the source of long history).
      2. Every `product_sales_*.csv` daily snapshot whose date is strictly
         AFTER the items file's end-date (closes the gap between the
         items export and "today" - the items export on the SFTP server
         is sometimes 1-3 days behind).

    Both file types share the same item-level schema (`order_item_id`,
    `created_at`, ...), so we can concat them raw and deduplicate by
    `order_item_id` before running standardisation. That avoids
    double-counting orders that appear in both an items snapshot and a
    daily snapshot.

    Items files smaller than `min_items_size_bytes` are skipped (this is
    how we tolerate the SFTP server occasionally producing a header-only
    cumulative file for the current day - see find_latest_file).

    Falls back gracefully:
      - If no items file is usable, returns the concatenation of all
        daily snapshots (still better than a single-day file).
      - If neither is available, raises FileNotFoundError.
    """
    directory = Path(directory)
    if not directory.exists():
        raise FileNotFoundError(f"Katalogen finns inte: {directory}")

    items_paths = glob.glob(str(directory / 'product_sales_items_*.csv'))
    daily_paths = [f for f in glob.glob(str(directory / 'product_sales_*.csv'))
                   if not os.path.basename(f).startswith('product_sales_items_')]

    def _extract_max_date(path):
        dates = re.findall(r'(\d{4}-\d{2}-\d{2})', os.path.basename(path))
        return max(dates) if dates else None

    items_with_size = []
    for p in items_paths:
        try:
            size = os.path.getsize(p)
        except OSError:
            size = 0
        items_with_size.append((p, size))
    usable_items = [(p, _extract_max_date(p)) for p, s in items_with_size
                    if s >= min_items_size_bytes and _extract_max_date(p)]
    usable_items.sort(key=lambda t: t[1], reverse=True)

    items_file = usable_items[0][0] if usable_items else None
    items_end_date = usable_items[0][1] if usable_items else None

    skipped = [(os.path.basename(p), s) for p, s in items_with_size
               if s < min_items_size_bytes]
    if skipped:
        print(
            f"  Hoppar över {len(skipped)} items-fil(er) under "
            f"{min_items_size_bytes} byte (sannolikt tomma exporter):"
        )
        for name, size in skipped[:5]:
            print(f"    - {name} ({size} byte)")
        if len(skipped) > 5:
            print(f"    ... och {len(skipped) - 5} till")

    selected_daily = []
    for p in daily_paths:
        d = _extract_max_date(p)
        if not d:
            continue
        if items_end_date is None or d > items_end_date:
            selected_daily.append((p, d))
    selected_daily.sort(key=lambda t: t[1])

    sources = []
    raw_dfs = []
    if items_file is not None:
        print(
            f"  Använder items-fil för historik: "
            f"{os.path.basename(items_file)} (t.o.m. {items_end_date})"
        )
        try:
            raw_dfs.append(pd.read_csv(str(items_file)))
            sources.append(items_file)
        except (OSError, pd.errors.EmptyDataError, pd.errors.ParserError) as e:
            print(f"  [WARNING] Kunde inte läsa items-fil: {e}")

    for p, d in selected_daily:
        try:
            raw_dfs.append(pd.read_csv(str(p)))
            sources.append(p)
            print(f"  Lägger till daglig snapshot: {os.path.basename(p)} ({d})")
        except (OSError, pd.errors.EmptyDataError, pd.errors.ParserError) as e:
            print(f"  [WARNING] Hoppar över {os.path.basename(p)}: {e}")

    if not raw_dfs:
        raise FileNotFoundError(
            f"Hittade inga användbara försäljningsfiler i {directory}. "
            f"Förväntade product_sales_items_*.csv (>= {min_items_size_bytes} byte) "
            f"eller product_sales_*.csv."
        )

    combined = pd.concat(raw_dfs, ignore_index=True, sort=False)

    if 'order_item_id' in combined.columns:
        before = len(combined)
        combined = combined.drop_duplicates(subset=['order_item_id'], keep='last')
        dropped = before - len(combined)
        if dropped:
            print(
                f"  Deduplicerade {dropped} rader på order_item_id "
                f"(overlap mellan items-fil och dagliga snapshots)"
            )
    else:
        # Belt-and-braces: if we somehow ended up with the period-aggregated
        # schema, dedupe on the natural key of that schema.
        natural_keys = [c for c in ('period', 'store_name', 'product_name')
                        if c in combined.columns]
        if len(natural_keys) >= 2:
            before = len(combined)
            combined = combined.drop_duplicates(subset=natural_keys, keep='last')
            dropped = before - len(combined)
            if dropped:
                print(f"  Deduplicerade {dropped} rader på {natural_keys}")

    print(f"  Kombinerade {len(sources)} fil(er) till {len(combined)} rader före standardisering")
    return _prepare_sales_df(combined, keep_unit=keep_unit)


# ---------------------------------------------------------------------------
# Forecasting — total over a forecast window (used by scripts 2 and 3)
# ---------------------------------------------------------------------------

# Hur mycket historik krävs innan vi släpper igenom ML-modellerna. Med
# < 2 år är säsongsmönstren för glesa för att en KNN/RF/GB-modell ska
# generalisera tillförlitligt; då används istället medel av de senaste
# kompletta veckorna (RECENT_WEEKS_FOR_AVG).
MIN_DAYS_FOR_ML = 730  # 2 år (2 * 365)

# Antal senaste kompletta kalenderveckor i baseline-snittet. 8 veckor ger
# mer stabila (och oftast högre) prognoser än 4 när historiken är kort,
# utan att dra in säsong från för länge sedan.
RECENT_WEEKS_FOR_AVG = 8

# Soft-vote-vikter: när produkten har minst 2 års historik blandar vi
# ML-prognosen med veckosnittet. Snittet får lite mer vikt än ML eftersom
# ML annars ofta undervärderar aktuell efterfrågan. Vikterna summerar till 1.0.
ML_VOTE_WEIGHT = 0.4
AVG_VOTE_WEIGHT = 0.6
assert abs((ML_VOTE_WEIGHT + AVG_VOTE_WEIGHT) - 1.0) < 1e-9


def _compute_mae(y_true, y_pred):
    """Mean absolute error that works even without sklearn.metrics."""
    return float(np.mean(np.abs(np.array(y_true) - np.array(y_pred))))


def _predict_weekly_average(daily_sales, forecast_days, ref_date,
                            n_weeks=RECENT_WEEKS_FOR_AVG):
    """
    Forecast using the mean of the latest *complete* calendar weeks.

    The current (possibly incomplete) week is excluded so the average is
    not pulled down by a partial week. Prefer weeks with actual sales when
    enough of them exist among the recent window (avoids underestimating
    after short stockouts that produced trailing zero-weeks).
    """
    ds = daily_sales.copy()
    ds['week'] = ds['date'].dt.to_period('W').apply(lambda r: r.start_time)
    weekly_sales = ds.groupby('week')['quantity'].sum().reset_index()
    weekly_sales = weekly_sales.sort_values('week')

    current_week_start = ref_date - pd.Timedelta(days=ref_date.dayofweek)
    complete_weeks = weekly_sales[weekly_sales['week'] < current_week_start]

    if len(complete_weeks) == 0:
        complete_weeks = weekly_sales

    if len(complete_weeks) == 0:
        return 0.0

    recent = complete_weeks.tail(min(n_weeks, len(complete_weeks)))
    active = recent[recent['quantity'] > 0]
    # Use active weeks if we have at least half the intended window;
    # otherwise keep all recent weeks (including zeros) so brand-new
    # products don't inflate from a single lucky week.
    if len(active) >= max(2, n_weeks // 2):
        avg_weekly = active['quantity'].mean()
    else:
        avg_weekly = recent['quantity'].mean()

    return max(0.0, float(avg_weekly / 7.0 * forecast_days))


def _predict_best_model(daily_sales, forecast_days, today):
    """
    Train several candidate models, evaluate each on a 28-day validation
    window, and use the winner to produce the final forecast. Falls back to
    a 4-week-average baseline if no ML model beats it.
    """
    ds = daily_sales.copy()
    ds['dayofweek'] = ds['date'].dt.dayofweek
    ds['month'] = ds['date'].dt.month
    ds['day'] = ds['date'].dt.day
    ds['weekofyear'] = ds['date'].dt.isocalendar().week.astype(int)
    ds['is_weekend'] = ds['dayofweek'].isin([5, 6]).astype(int)

    feature_cols = ['dayofweek', 'month', 'day', 'weekofyear', 'is_weekend']
    X = ds[feature_cols]
    y = ds['quantity']

    val_days = min(28, len(ds) // 4)
    if val_days < 7:
        return _predict_weekly_average(daily_sales, forecast_days, today)

    X_train, X_val = X.iloc[:-val_days], X.iloc[-val_days:]
    y_train, y_val = y.iloc[:-val_days], y.iloc[-val_days:]

    train_end = ds['date'].iloc[-val_days - 1]
    baseline_total = _predict_weekly_average(
        ds[['date', 'quantity']].iloc[:-val_days].copy(), val_days, train_end
    )
    baseline_daily = baseline_total / val_days if val_days > 0 else 0.0
    baseline_mae = _compute_mae(y_val, [baseline_daily] * len(y_val))

    candidates = {
        'knn_3': KNeighborsRegressor(n_neighbors=3),
        'knn_5': KNeighborsRegressor(n_neighbors=5),
        'knn_7': KNeighborsRegressor(n_neighbors=7),
    }
    if _ADVANCED_ML:
        candidates['linear'] = LinearRegression()
        candidates['ridge'] = Ridge(alpha=1.0)
        candidates['rf'] = RandomForestRegressor(
            n_estimators=50, max_depth=10, random_state=42)
        candidates['gb'] = GradientBoostingRegressor(
            n_estimators=50, max_depth=5, random_state=42)

    best_mae = baseline_mae
    best_model = None
    for _, model in candidates.items():
        try:
            model.fit(X_train, y_train)
            preds = np.maximum(model.predict(X_val), 0)
            mae = _compute_mae(y_val, preds)
            if mae < best_mae:
                best_mae = mae
                best_model = model
        except Exception:
            continue

    if best_model is None:
        return _predict_weekly_average(daily_sales, forecast_days, today)

    best_model.fit(X, y)

    tomorrow = today + pd.Timedelta(days=1)
    future_dates = pd.date_range(tomorrow, periods=forecast_days, freq='D')
    fdf = pd.DataFrame({'date': future_dates})
    fdf['dayofweek'] = fdf['date'].dt.dayofweek
    fdf['month'] = fdf['date'].dt.month
    fdf['day'] = fdf['date'].dt.day
    fdf['weekofyear'] = fdf['date'].dt.isocalendar().week.astype(int)
    fdf['is_weekend'] = fdf['dayofweek'].isin([5, 6]).astype(int)

    y_future = np.maximum(best_model.predict(fdf[feature_cols]), 0)
    return float(y_future.sum())


def predict_product_sales(product_df, forecast_days):
    """
    Total expected sales for `forecast_days` days, based on historical data.

      - < 2 years of history  → average of the latest RECENT_WEEKS_FOR_AVG
        complete weeks (recency-biased baseline; prefers weeks with sales).
      - >= 2 years of history → soft-vote between the best ML model
        (KNN/Linear/Ridge/RF/GB) and the same recent-weeks baseline.
    """
    daily_sales = product_df.groupby('date')['quantity'].sum().reset_index()
    daily_sales = daily_sales.sort_values('date')
    daily_sales['date'] = pd.to_datetime(daily_sales['date'])
    daily_sales['quantity'] = daily_sales['quantity'].fillna(0)

    today = pd.Timestamp.now().normalize()
    first_date = daily_sales['date'].min()

    date_range = pd.date_range(start=first_date, end=today, freq='D')
    daily_sales = daily_sales.set_index('date').reindex(
        date_range, fill_value=0).reset_index()
    daily_sales = daily_sales.rename(columns={'index': 'date'})
    daily_sales['quantity'] = daily_sales['quantity'].fillna(0)

    avg_pred = _predict_weekly_average(daily_sales, forecast_days, today)
    if len(daily_sales) < MIN_DAYS_FOR_ML:
        return avg_pred

    ml_pred = _predict_best_model(daily_sales, forecast_days, today)
    return ML_VOTE_WEIGHT * ml_pred + AVG_VOTE_WEIGHT * avg_pred


# ---------------------------------------------------------------------------
# Forecasting — week-by-week (used by script 4 for the graphical picking list)
# ---------------------------------------------------------------------------

def _empty_forecast_df(first_future_monday, future_weeks):
    """Returns a forecast DataFrame filled with zeros."""
    rows = []
    for i in range(future_weeks):
        ws = first_future_monday + pd.Timedelta(weeks=i)
        iso = ws.isocalendar()
        rows.append({
            'week_start': ws,
            'quantity': 0.0,
            'week_number': iso.week,
            'year': iso.year,
            'week_year': f"{iso.year}-W{int(iso.week):02d}",
        })
    return pd.DataFrame(rows)


def _forecast_weeks_short_history(daily_sales, first_future_monday, future_weeks, ref_date):
    """For < 5 months history: KNN on weekly-aggregated data."""
    ds = daily_sales.copy()
    ds['week'] = ds['date'].dt.to_period('W').apply(lambda r: r.start_time)
    weekly = ds.groupby('week')['quantity'].sum().reset_index()
    weekly = weekly.sort_values('week')

    current_week_start = ref_date - pd.Timedelta(days=ref_date.dayofweek)
    complete = weekly[weekly['week'] < current_week_start].copy()

    if len(complete) == 0:
        complete = weekly.copy()
    if len(complete) == 0:
        return _empty_forecast_df(first_future_monday, future_weeks)

    complete['week_of_year'] = complete['week'].dt.isocalendar().week.astype(int)
    complete['month'] = complete['week'].dt.month

    k = min(4, len(complete))
    model = KNeighborsRegressor(n_neighbors=k, weights='distance')
    model.fit(complete[['week_of_year', 'month']], complete['quantity'])

    rows = []
    for i in range(future_weeks):
        ws = first_future_monday + pd.Timedelta(weeks=i)
        iso = ws.isocalendar()
        pred = max(0.0, float(model.predict([[int(iso.week), ws.month]])[0]))
        rows.append({
            'week_start': ws,
            'quantity': pred,
            'week_number': iso.week,
            'year': iso.year,
            'week_year': f"{iso.year}-W{int(iso.week):02d}",
        })
    return pd.DataFrame(rows)


def _forecast_weeks_best_model(daily_sales, first_future_monday, future_weeks, today):
    """For >= 5 months history: best of KNN/Linear/Ridge/RF/GB, day-by-day, aggregated to weekly."""
    ds = daily_sales.copy()
    ds['dayofweek'] = ds['date'].dt.dayofweek
    ds['month'] = ds['date'].dt.month
    ds['day'] = ds['date'].dt.day
    ds['weekofyear'] = ds['date'].dt.isocalendar().week.astype(int)
    ds['is_weekend'] = ds['dayofweek'].isin([5, 6]).astype(int)

    feature_cols = ['dayofweek', 'month', 'day', 'weekofyear', 'is_weekend']
    X = ds[feature_cols]
    y = ds['quantity']

    val_days = min(28, len(ds) // 4)
    if val_days < 7:
        return _uniform_weekly_avg_forecast(
            daily_sales, first_future_monday, future_weeks, today
        )

    X_train, X_val = X.iloc[:-val_days], X.iloc[-val_days:]
    y_train, y_val = y.iloc[:-val_days], y.iloc[-val_days:]

    train_ds = ds.iloc[:-val_days].copy()
    train_ds['week'] = train_ds['date'].dt.to_period('W').apply(lambda r: r.start_time)
    train_weekly = train_ds.groupby('week')['quantity'].sum()
    train_end = train_ds['date'].iloc[-1]
    cws = train_end - pd.Timedelta(days=train_end.dayofweek)
    complete_train_weeks = train_weekly[train_weekly.index < cws]
    if len(complete_train_weeks) > 0:
        baseline_daily = complete_train_weeks.tail(
            min(RECENT_WEEKS_FOR_AVG, len(complete_train_weeks))).mean() / 7.0
    elif len(train_weekly) > 0:
        baseline_daily = train_weekly.mean() / 7.0
    else:
        baseline_daily = 0.0
    baseline_mae = _compute_mae(y_val, [baseline_daily] * len(y_val))

    candidates = {
        'knn_3': KNeighborsRegressor(n_neighbors=3),
        'knn_5': KNeighborsRegressor(n_neighbors=5),
        'knn_7': KNeighborsRegressor(n_neighbors=7),
    }
    if _ADVANCED_ML:
        candidates['linear'] = LinearRegression()
        candidates['ridge'] = Ridge(alpha=1.0)
        candidates['rf'] = RandomForestRegressor(
            n_estimators=50, max_depth=10, random_state=42)
        candidates['gb'] = GradientBoostingRegressor(
            n_estimators=50, max_depth=5, random_state=42)

    best_mae = baseline_mae
    best_model = None
    for _, model in candidates.items():
        try:
            model.fit(X_train, y_train)
            preds = np.maximum(model.predict(X_val), 0)
            mae = _compute_mae(y_val, preds)
            if mae < best_mae:
                best_mae = mae
                best_model = model
        except Exception:
            continue

    if best_model is None:
        return _uniform_weekly_avg_forecast(
            daily_sales, first_future_monday, future_weeks, today
        )

    best_model.fit(X, y)

    total_days = future_weeks * 7
    future_dates = pd.date_range(
        first_future_monday, periods=total_days, freq='D')
    fdf = pd.DataFrame({'date': future_dates})
    fdf['dayofweek'] = fdf['date'].dt.dayofweek
    fdf['month'] = fdf['date'].dt.month
    fdf['day'] = fdf['date'].dt.day
    fdf['weekofyear'] = fdf['date'].dt.isocalendar().week.astype(int)
    fdf['is_weekend'] = fdf['dayofweek'].isin([5, 6]).astype(int)
    fdf['quantity'] = np.maximum(best_model.predict(fdf[feature_cols]), 0)

    fdf['week_start'] = fdf['date'] - pd.to_timedelta(
        fdf['date'].dt.dayofweek, unit='D')
    weekly = fdf.groupby('week_start')['quantity'].sum().reset_index()
    weekly = weekly.sort_values('week_start')
    weekly['week_number'] = weekly['week_start'].dt.isocalendar().week
    weekly['year'] = weekly['week_start'].dt.isocalendar().year
    weekly['week_year'] = weekly.apply(
        lambda r: f"{int(r['year'])}-W{int(r['week_number']):02d}", axis=1
    )
    return weekly


def _uniform_weekly_avg_forecast(daily_sales, first_future_monday,
                                 future_weeks, ref_date):
    """
    Producera en vecka-för-vecka-prognos där varje vecka har samma
    värde - medel från senaste kompletta veckorna (RECENT_WEEKS_FOR_AVG).
    Används som baseline i soft-vote mot ML-prognosen, och som enda
    prognos när historiken är kortare än 2 år.
    """
    week_total = _predict_weekly_average(daily_sales, 7, ref_date)
    rows = []
    for i in range(future_weeks):
        ws = first_future_monday + pd.Timedelta(weeks=i)
        iso = ws.isocalendar()
        rows.append({
            'week_start': ws,
            'quantity': float(week_total),
            'week_number': int(iso.week),
            'year': int(iso.year),
            'week_year': f"{int(iso.year)}-W{int(iso.week):02d}",
        })
    return pd.DataFrame(rows)


def predict_weekly_sales(product_df, future_weeks=4):
    """
    Week-by-week sales forecast for the next `future_weeks` weeks.
    Returns DataFrame with: week_start, quantity, week_number, year, week_year.

      - < 2 years history  → recent-weeks average (same value each week).
      - >= 2 years history → soft-vote per week between best ML model
        and recent-weeks-average baseline.
    """
    daily_sales = product_df.groupby('date')['quantity'].sum().reset_index()
    daily_sales = daily_sales.sort_values('date')
    daily_sales['date'] = pd.to_datetime(daily_sales['date'])
    daily_sales['quantity'] = daily_sales['quantity'].fillna(0)

    today = pd.Timestamp.now().normalize()
    first_date = daily_sales['date'].min()

    date_range = pd.date_range(start=first_date, end=today, freq='D')
    daily_sales = daily_sales.set_index('date').reindex(
        date_range, fill_value=0).reset_index()
    daily_sales = daily_sales.rename(columns={'index': 'date'})
    daily_sales['quantity'] = daily_sales['quantity'].fillna(0)

    tomorrow = today + pd.Timedelta(days=1)
    days_until_monday = (7 - tomorrow.weekday()) % 7
    first_future_monday = tomorrow if days_until_monday == 0 else tomorrow + \
        pd.Timedelta(days=days_until_monday)

    avg_weekly = _uniform_weekly_avg_forecast(
        daily_sales, first_future_monday, future_weeks, today
    )
    if len(daily_sales) < MIN_DAYS_FOR_ML:
        return avg_weekly

    ml_weekly = _forecast_weeks_best_model(
        daily_sales, first_future_monday, future_weeks, today
    )

    # Soft-vote per vecka: behåller ML-modellens vecko-shape men dras
    # mot recent-weeks-snittet så framtida veckor inte undervärderas.
    merged = ml_weekly.merge(
        avg_weekly[['week_start', 'quantity']],
        on='week_start', suffixes=('', '_avg'), how='left',
    )
    merged['quantity_avg'] = merged['quantity_avg'].fillna(0.0)
    merged['quantity'] = (
        ML_VOTE_WEIGHT * merged['quantity']
        + AVG_VOTE_WEIGHT * merged['quantity_avg']
    )
    return merged.drop(columns=['quantity_avg'])
